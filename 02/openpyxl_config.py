import sys
import datetime

from telnetlib import Telnet
from openpyxl import Workbook


def func():
    # コマンドライン引数からIPアドレス、ユーザー、パスワードを取得 ...(1)
    host = sys.argv[1]
    username = sys.argv[2]
    password = sys.argv[3]

    # 作成日時を保持                                           ...(2)
    now = datetime.datetime.now()

    # Telnetインスタンスを作成しログイン                        ...(3)
    print(host)
    tlnet = Telnet(host)
    tlnet.read_until(b"Password:")
    tlnet.write(b"\n")
    tlnet.read_until(b"Username:")
    tlnet.write(username.encode("ascii") + b"\n")
    tlnet.read_until(b"Password:")
    tlnet.write(password.encode("ascii") + b"\n")
    tlnet.read_until(b">")

    # コンソールの表示を英語、文字コードをASCIIにする            ...(4)
    tlnet.write(b"console character en.ascii\n")
    tlnet.read_until(b">")

    # コンソールのスクロールを止めない
    tlnet.write(b"console lines infinity\n")
    tlnet.read_until(b">")

    # コンソールの1行あたりの表示文字数を200にする
    tlnet.write(b"console columns 200\n")
    tlnet.read_until(b">")

    # Excelファイルのオブジェクトを作成                        ...(5)
    wb = Workbook()
    ws = wb.active
    ws.title = host

    # 作成日時を出力
    row = 1
    ws.cell(row, 1).value = "作成日時"
    ws.cell(row, 2).value = now.strftime('%Y/%m/%d %H:%M:%S')

    # 機器情報を取得                                        ...(6)
    row = row + 1
    tlnet.write(b"show environment\n")
    env = tlnet.read_until(b">").decode("utf-8")
    envlines = env.splitlines()
    for envline in envlines:
        if "Rev." in envline:
            row = row + 1
            ws.cell(row, 1).value = "機種名"
            ws.cell(row, 2).value = envline

        elif "main:" in envline:
            target1 = "serial="
            index1 = envline.find(target1)
            target2 = "MAC-Address="
            index2 = envline.find(target2)
            serial = envline[index1+len(target1):index2]
            row = row + 1
            ws.cell(row, 1).value = "シリアル番号"
            ws.cell(row, 2).value = serial

        elif "Boot time" in envline:
            row = row + 1
            ws.cell(row, 1).value = "起動時刻"
            ws.cell(row, 2).value = envline.split(": ")[1]

        elif "Elapsed time from boot" in envline:
            row = row + 1
            ws.cell(row, 1).value = "起動からの経過時間"
            ws.cell(row, 2).value = envline.split(": ")[1]

        elif "Inside Temperature" in envline:
            row = row + 1
            ws.cell(row, 1).value = "筐体内温度(℃)"
            ws.cell(row, 2).value = envline.split(": ")[1]

    # コンフィグを取得                                      ...(7)
    row = row + 2
    ws.cell(row, 1).value = "コンフィグ"
    cnfout = False
    tlnet.write(b"show config\n")
    cnf = tlnet.read_until(b">").decode("utf-8")
    cnflines = cnf.splitlines()
    for cnfline in cnflines:
        # 「ip route default」が出現したら、以降すべてを出力
        if "ip route default" in cnfline:
            cnfout = True
        # 「>」が出現したら出力終了
        elif ">" in cnfline:
            break

        if cnfout == True:
            row = row + 1
            ws.cell(row, 1).value = cnfline

    # Telnet接続を閉じる                                   ...(8)
    tlnet.close()

    # Excelファイルを保存                                  ...(9)
    wb.save(host + '_' + now.strftime('%Y%m%d%H%M%S') + '.xlsx')


if __name__ == "__main__":                                      # ... 8
    func()
